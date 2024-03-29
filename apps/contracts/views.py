import datetime
from django.utils.timezone import is_aware, make_naive
from django.http import HttpResponse
from openpyxl import Workbook
from django.http import HttpResponseBadRequest
from rest_framework.permissions import AllowAny  # Importar AllowAny
from rest_framework.pagination import PageNumberPagination
from rest_framework.pagination import LimitOffsetPagination
from django.shortcuts import render
# Create your views here.


# Rest Framework imports
from rest_framework import viewsets
from rest_framework import permissions
from rest_framework import generics
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.views import APIView

# models
from apps.contracts.models import (
    Contratos, OrdenDeTrabajo,
    Lotes, Satelites)

from apps.user.models import User

# serializers
from apps.contracts.serializers import (
    ContratosSerializer, OrdenDeTrabajoSerializer,
    LotesSerializer, PersonalSerializer, SatelitesSerializer
)


class ListContrats(viewsets.ModelViewSet):
    """
    List all contracts
    """

    def get_queryset(self):

        queryseet = super().get_queryset()

        estado_lotes = self.request.query_params.get('estado', None)
        if estado_lotes is not None:
            queryseet = queryseet.filter(estado=estado_lotes)

        return queryseet

    permission_classes = [AllowAny]  # Permitir cualquier persona

    queryset = Contratos.objects.all()
    serializer_class = ContratosSerializer
    permission_classes = [permissions.IsAuthenticated]


class OrdenOfWorkView(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = OrdenDeTrabajo.objects.all()
    serializer_class = OrdenDeTrabajoSerializer
    # permission_classes = [permissions.IsAuthenticated]


class PersonasView(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = User.objects.all()
    serializer_class = PersonalSerializer
    permission_classes = [AllowAny]  # Permitir cualquier persona

    def get_queryset(self):
        cargo = self.request.query_params.get('cargo', None)
        if cargo is not None:
            return User.objects.filter(cargo=cargo)
        return super().get_queryset()

    # permission_classes = [permissions.IsAuthenticated]


class LotView(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """

    def get_queryset(self):

        queryseet = super().get_queryset()

        estado_lotes = self.request.query_params.get('estado', None)
        if estado_lotes is not None:
            queryseet = queryseet.filter(estado=estado_lotes)

        return queryseet

    permission_classes = [AllowAny]  # Permitir cualquier persona
    queryset = Lotes.objects.all()
    serializer_class = LotesSerializer
    # permission_classes = [permissions.IsAuthenticated]


class PersonasMeView(APIView):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = User.objects.all()
    serializer_class = PersonalSerializer
    permission_classes = [AllowAny]  # Permitir cualquier persona

    def get(self, request, format=None):
        user = request.user
        print(user)
        serializer_data = PersonalSerializer(user).data
        # data = {
        #     'id': user.id,
        #     'username': user.username,
        #     'first_name': user.first_name,
        #     'last_name': user.area,
        #     'cargo': user.cargo,
        # }
        return Response(serializer_data, status=status.HTTP_200_OK)


class OrdenDeTrabajoPorContrato(APIView):
    """
    API endpoint that allows users to be viewed or edited.
    """

    permission_classes = [AllowAny]  # Permitir cualquier persona

    def get(self, request, format=None, *args, **kwargs):

        id = kwargs['id']
        contrato = Contratos.objects.filter(id=id).first()

        if not contrato:
            return Response({'detail': 'No existe el contrato'},
                            status=status.HTTP_400_BAD_REQUEST)

        ordenes_de_trabajo = OrdenDeTrabajo.objects.filter(
            numero_de_contrato=contrato.numero_contrato)

        # Paginar los resultados
        paginator = PageNumberPagination()
        paginated_ordenes = paginator.paginate_queryset(
            ordenes_de_trabajo, request)

        data = OrdenDeTrabajoSerializer(paginated_ordenes, many=True).data

        return paginator.get_paginated_response(data)


class LotesDeTrabajoPorContrato(APIView):
    """
    API endpoint that allows users to be viewed or edited.
    """
    permission_classes = [AllowAny]  # Permitir cualquier persona

    def get(self, request, format=None, *args, **kwargs):
        id = kwargs['id']
        contrato = Contratos.objects.filter(id=id).first()

        if not contrato:
            return Response({'detail': 'No existe el contrato'},
                            status=status.HTTP_400_BAD_REQUEST)

        lotes_de_trabajo = Lotes.objects.filter(
            numero_de_contrato=contrato.numero_contrato)

        # get the number of the contract of query params
        estado_lotes = request.query_params.get('estado', None)
        if estado_lotes is not None:
            lotes_de_trabajo = lotes_de_trabajo.filter(estado=estado_lotes)

        # Paginar los resultados
        paginator = PageNumberPagination()
        paginated_lotes = paginator.paginate_queryset(
            lotes_de_trabajo, request)

        data = LotesSerializer(paginated_lotes, many=True).data

        return paginator.get_paginated_response(data)


class LotesDeTrabajoPorOrdenDeTrabajo(APIView):
    """
    API endpoint that allows users to be viewed or edited.
    """
    permission_classes = [AllowAny]  # Permitir cualquier persona

    def get(self, request, format=None, *args, **kwargs):
        id = kwargs['id']
        orden_de_trabajo = OrdenDeTrabajo.objects.filter(id=id).first()

        if not orden_de_trabajo:
            return Response({'detail': 'No existe el orden de trabajo'},
                            status=status.HTTP_400_BAD_REQUEST)

        lotes_de_trabajo = Lotes.objects.filter(
            numero_de_orden=orden_de_trabajo.numero_de_orden)

        # get the number of the contract of query params
        estado_lotes = request.query_params.get('estado', None)
        if estado_lotes is not None:
            lotes_de_trabajo = lotes_de_trabajo.filter(estado=estado_lotes)

        # Paginar los resultados
        paginator = PageNumberPagination()
        paginated_lotes = paginator.paginate_queryset(
            lotes_de_trabajo, request)

        data = LotesSerializer(paginated_lotes, many=True).data

        return paginator.get_paginated_response(data)


class LotesDeTrabajoAsignado(APIView):
    """
    API endpoint that allows users to be viewed or edited.
    """
    permission_classes = [AllowAny]  # Permitir cualquier persona

    def get(self, request, format=None, *args, **kwargs):
        area = kwargs['area']
        lotes_de_trabajo = Lotes.objects.filter(
            asignado__icontains=area)

        # get the number of the contract of query params
        estado_lotes = request.query_params.get('estado', None)
        if estado_lotes is not None:
            lotes_de_trabajo = lotes_de_trabajo.filter(estado=estado_lotes)
        else:
            lotes_de_trabajo = lotes_de_trabajo.exclude(estado="Finalizado")
        # Paginar los resultados
        paginator = PageNumberPagination()
        paginated_lotes = paginator.paginate_queryset(
            lotes_de_trabajo, request)

        data = LotesSerializer(paginated_lotes, many=True).data

        return paginator.get_paginated_response(data)


class SatelitesView(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Satelites.objects.all()
    serializer_class = SatelitesSerializer
    permission_classes = [AllowAny,]  # Permitir cualquier persona


class ExcelGeneratorView(APIView):

    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Satelites.objects.all()
    serializer_class = SatelitesSerializer
    permission_classes = [AllowAny,]  # Permitir cualquier persona


# import HttpResponse


class ExcelGeneratorView(APIView):
    permission_classes = [AllowAny,]  # Permitir cualquier persona

    def get(self, request, *args, **kwargs):
        models = request.GET.getlist('models', [])

        if not models:
            return Response({"error": "No se proporcionaron modelos."}, status=400)

        wb = Workbook()

        try:
            for model_name in models:
                model = Contratos

                w1 = wb.create_sheet(title=model_name)

                field_names = [field.name for field in model._meta.fields]
                w1.append(field_names)
                for obj in model.objects.all():
                    row = []
                    for field in obj._meta.fields:
                        value = getattr(obj, field.name)
                        if isinstance(value, datetime.datetime):
                            if is_aware(value):
                                value = make_naive(value)
                        row.append(value)
                    w1.append(row)

                w2 = wb.create_sheet(title=model_name)

                model = Lotes
                field_names = [field.name for field in model._meta.fields]
                w2.append(field_names)
                for obj in model.objects.all():
                    row = []
                    for field in obj._meta.fields:
                        value = getattr(obj, field.name)
                        if isinstance(value, datetime.datetime):
                            if is_aware(value):
                                value = make_naive(value)
                        row.append(value)
                    w2.append(row)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="data.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=400)
